from django.http import HttpResponse
import sys 
sys.path.append('/usr/lib/python3/dist-packages')
import openpyxl
from django.shortcuts import redirect, render
from django.contrib.auth import authenticate, login, logout
from django.contrib import  messages
from django.contrib.auth.decorators import login_required
from database.models import *

def signin(request):
    if request.user.is_authenticated:
        logout(request)
        messages.success(request, "You are Logged out")
        return render(request, "Login.html")

    if request.method=="POST":
        ID=(request.POST.get('RollN'))
        Pass=(request.POST.get('password'))
        
        user = authenticate(username=ID, password=Pass)  
 
        if user is not None:
            login(request, user)
            user = Student.objects.get(roll_num=request.user.id)
            info = {'roll_num': user.roll_num, 'name':str(user.Firstname),'name2': str(user.Lastname), 'mail':user.Email, 'number':user.Number, 'address':user.Address}
            return redirect('home/', info)

        else:
            messages.error(request, "Invalid username or password")

    return render(request, 'Login.html')

@login_required
def home(request):
    user = Student.objects.get(roll_num=request.user.id)
    info = {'roll_num': user.roll_num, 'name':str(user.Firstname),'name2': str(user.Lastname), 'mail':user.Email, 'number':user.Number, 'address':user.Address}

    if request.method=='POST':
        n7=str(request.POST.get('p2'))
        n8=str(request.POST.get('p3'))
        if(n7!=n8):
            messages.info(request, 'New passwords entered does not match')
            return render(request, "index.html", info)

        else:
            messages.info(request, 'Password Changed successfully')
            u = User.objects.get(username=request.user.username)
            u.set_password(request.POST['p2'])
            u.save()

    return render(request, "index.html", info)   

@login_required
def accesscoupon(request):
    u = Student.objects.get(roll_num=request.user.id)
    if (Coupon.objects.filter(Student=u).exists()):
        coupons= Coupon.objects.filter(Student=u).values()
        context={'coupons': coupons}
        return render(request, "Access_Coupon.html", context)  
    else:
        coupons=[{'Date': '-', 'Breakfast':'-', 'Lunch':'-', 'Evening_snacks':'-', 'Dinner':'-'}]
        messages.info(request, 'No coupons bought')
        context={'coupons': coupons} 
        return render(request, "Access_Coupon.html", context)    


@login_required
def buycoupon(request):
    TotalAmt=0
    n1=0
    n2=0
    n3=0
    n4=0
    data={
            'TotalAmt':TotalAmt,
            'n1':n1,
            'n2':n2,
            'n3':n3,
            'n4':n4,
        }

    try:  
        n1=int(request.GET.get('n1'))
        n2=int(request.GET.get('n2'))
        n3=int(request.GET.get('n3'))
        n4=int(request.GET.get('n4'))
        n5=(request.GET.get('n5'))
        TotalAmt= n1*30+n2*50+n3*20+n4*50
        data={
            'TotalAmt':TotalAmt,
            'n1':n1,
            'n2':n2,
            'n3':n3,
            'n4':n4,
            'n5':n5
        }
        u = Student.objects.get(roll_num=request.user.id)
        en=Coupon(Student=u, Breakfast=n1, Lunch=n2, Evening_snacks=n3, Dinner=n4, Date=n5)
        en.save()   
    except:
        pass    

    return render(request, "Buy_Coupon.html",data)

@login_required
def complaintstatus(request):
    u = Student.objects.get(roll_num=request.user.id)
    if(ComplaintRecord.objects.filter(Student=u).exists()):
        complaints= (ComplaintRecord.objects.filter(Student=u).values())
        context={'complaints': complaints}
    else:
        messages.info(request, 'No complaints registered')
        complaints= [{'Complaint_ID': '-', 'Subject':'-', 'Description':'-', 'Status':'-', 'Comments':'-'}]
        context={'complaints': complaints}
    return render(request, "complaint_status.html", context)   


def messmenu(request):
    return render(request, "mess_menu.html")

@login_required
def menumon(request):
    wb = openpyxl.load_workbook('media/mess_menu/Mess_Menu.xlsx')
    worksheet = wb['Sheet1']
    list={'Day':worksheet["A2"].value, 'Breakfast':worksheet["B2"].value, 'Lunch':worksheet["C2"].value, 'Evening_snacks':worksheet["D2"].value, 'Dinner':worksheet["E2"].value}
    print(list)    
    return render(request, "mess_menu_mon.html", list)

@login_required
def menutue(request):
    wb = openpyxl.load_workbook('media/mess_menu/Mess_Menu.xlsx')
    worksheet = wb['Sheet1']
    list={'Day':worksheet["A3"].value, 'Breakfast':worksheet["B3"].value, 'Lunch':worksheet["C3"].value, 'Evening_snacks':worksheet["D3"].value, 'Dinner':worksheet["E3"].value}
        
    return render(request, "mess_menu_mon.html", list)

@login_required
def menuwed(request):
    wb = openpyxl.load_workbook('media/mess_menu/Mess_Menu.xlsx')
    worksheet = wb['Sheet1']
    list={'Day':worksheet["A4"].value, 'Breakfast':worksheet["B4"].value, 'Lunch':worksheet["C4"].value, 'Evening_snacks':worksheet["D4"].value, 'Dinner':worksheet["E4"].value}
        
    return render(request, "mess_menu_mon.html", list)

@login_required
def menuthu(request):
    wb = openpyxl.load_workbook('media/mess_menu/Mess_Menu.xlsx')
    worksheet = wb['Sheet1']
    list={'Day':worksheet["A5"].value, 'Breakfast':worksheet["B5"].value, 'Lunch':worksheet["C5"].value, 'Evening_snacks':worksheet["D5"].value, 'Dinner':worksheet["E5"].value}
        
    return render(request, "mess_menu_mon.html", list)

@login_required
def menufri(request):
    wb = openpyxl.load_workbook('media/mess_menu/Mess_Menu.xlsx')
    worksheet = wb['Sheet1']
    list={'Day':worksheet["A6"].value, 'Breakfast':worksheet["B6"].value, 'Lunch':worksheet["C6"].value, 'Evening_snacks':worksheet["D6"].value, 'Dinner':worksheet["E6"].value}
        
    return render(request, "mess_menu_mon.html", list)    

@login_required
def menusat(request):
    wb = openpyxl.load_workbook('media/mess_menu/Mess_Menu.xlsx')
    worksheet = wb['Sheet1']
    list={'Day':worksheet["A7"].value, 'Breakfast':worksheet["B7"].value, 'Lunch':worksheet["C7"].value, 'Evening_snacks':worksheet["D7"].value, 'Dinner':worksheet["E7"].value}
        
    return render(request, "mess_menu_mon.html", list)

@login_required
def menusun(request):
    wb = openpyxl.load_workbook('media/mess_menu/Mess_Menu.xlsx')
    worksheet = wb['Sheet1']
    list={'Day':worksheet["A8"].value, 'Breakfast':worksheet["B8"].value, 'Lunch':worksheet["C8"].value, 'Evening_snacks':worksheet["D8"].value, 'Dinner':worksheet["E8"].value}
        
    return render(request, "mess_menu_mon.html", list)

@login_required
def messschedule(request):
    u=str(request.user.username)
    v=u[0:2]
    if (v=='19' or v=='21'):
        a='Mess A'
    else:
        a='Mess B'   

    return render(request, "mess_schedule.html", {'v':a})

@login_required
def registercomplaint(request):

    if request.method=='POST':
        n1=request.POST.get('sub')
        n2=request.POST.get('msg')
            #Add it to database
        print (request.POST.get('msg'))
        u = Student.objects.get(roll_num=request.user.id)
        en=ComplaintRecord(Student=u, Subject=n1, Description=n2)
        en.save()
    return render(request, "register_complaint.html")



